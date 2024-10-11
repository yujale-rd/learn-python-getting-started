# Python内部未提供处理Excel文件的功能，想要在Python中操作Excel需要按照第三方的模块。
#  openpyxl：用于处理 .xlsx 格式的 Excel 文件，支持读取和写入
#  xlsxwriter：专门用于写入 .xlsx 文件，支持格式化、图表等功能。	。

# 1. 使用 openpyxl 处理 Excel 文件
#
# openpyxl 是用于处理 Excel .xlsx 文件的最流行的库之一。它支持读取和写入工作表，还可以进行单元格的格式化操作。
#
# 安装 openpyxl
#
# 首先，你需要安装 openpyxl：
# pip install openpyxl

# 读取 Excel 文件
#
# 使用 openpyxl 读取 Excel 文件的基本步骤如下：

import openpyxl

# 打开 Excel 文件
workbook = openpyxl.load_workbook('example.xlsx')

# 选择一个工作表
sheet = workbook.active  # 默认选择活动的工作表
# 或者你可以指定工作表的名称
# sheet = workbook['Sheet1']

# 读取特定单元格的数据
cell_value = sheet['A1'].value
print(f"A1 cell value: {cell_value}")

# 遍历所有的行和列
for row in sheet.iter_rows(values_only=True):
    print(row)

# 写入 Excel 文件
import openpyxl

#

# 创建一个新的 Excel 工作簿
workbook = openpyxl.Workbook()

# 选择活动的工作表
sheet = workbook.active

# 写入数据到指定单元格
sheet['A1'] = 'Name'
sheet['B1'] = 'Age'

# 写入多行数据
data = [
    ['Alice', 30],
    ['Bob', 25],
    ['Charlie', 35]
]
for row in data:
    sheet.append(row)

# 保存工作簿
workbook.save('output.xlsx')

# 修改现有 Excel 文件

import openpyxl

# 打开已有的 Excel 文件
workbook = openpyxl.load_workbook('example.xlsx')
sheet = workbook.active

# 修改单元格的值
sheet['A1'] = 'Modified Name'

# 保存修改后的文件
workbook.save('example_modified.xlsx')

# 使用 xlsxwriter 进行高级写入
#
# 如果需要进行更高级的 Excel 文件操作，比如添加图表、格式化单元格等，可以使用 xlsxwriter。

# pip install xlsxwriter

# import xlsxwriter
#
# # 创建一个新的 Excel 文件和工作表
# workbook = xlsxwriter.Workbook('chart_example.xlsx')
# worksheet = workbook.add_worksheet()
#
# # 写入数据
# data = [10, 40, 50, 20, 10, 50]
# worksheet.write_column('A1', data)
#
# # 创建图表对象
# chart = workbook.add_chart({'type': 'column'})
#
# # 配置图表数据来源
# chart.add_series({'values': '=Sheet1!$A$1:$A$6'})
#
# # 插入图表到工作表中
# worksheet.insert_chart('C1', chart)
#
# # 关闭文件（保存）
# workbook.close()