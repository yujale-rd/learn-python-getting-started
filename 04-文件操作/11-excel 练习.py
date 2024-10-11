# 任务：给定一个 Excel 文件 data.xlsx，文件中包含多个工作表。请编写 Python 程序，读取第一个工作表中的所有数据，并打印出来。
#
# 思路：
#
# 	•	使用 openpyxl 或 pandas 来读取文件。
# 	•	获取活动工作表或第一个工作表。
# 	•	使用循环或其他方法遍历所有行，打印每行内容。

import openpyxl

# 打开 Excel 文件
workbook = openpyxl.load_workbook('data.xlsx')

# 获取第一个工作表
sheet = workbook.active

# 遍历工作表中的所有行并打印
for row in sheet.iter_rows(values_only=True):
    print(row)

# 任务：给定一个列表，列表中包含一些人的姓名和年龄，如 [['Alice', 30], ['Bob', 25], ['Charlie', 35]]。将这些数据写入新的 Excel 文件 people.xlsx 中，并将列标题设置为 “Name” 和 “Age”。
#
# 思路：
#
# 	•	使用 openpyxl 创建新文件。
# 	•	设置第一行标题。
# 	•	使用 append() 方法将列表数据添加到 Excel 表格中。

import openpyxl

# 创建新的 Excel 工作簿
workbook = openpyxl.Workbook()

# 获取活动的工作表
sheet = workbook.active

# 写入列标题
sheet.append(['Name', 'Age'])

# 数据列表
data = [['Alice', 30], ['Bob', 25], ['Charlie', 35]]

# 写入数据
for person in data:
    sheet.append(person)

# 保存文件
workbook.save('people.xlsx')

# 任务：给定一个 Excel 文件 scores.xlsx，其中包含学生的考试成绩。请读取这些成绩，计算所有学生的平均成绩并输出。
#
# 思路：
#
# 	•	使用 pandas 读取文件。
# 	•	通过列名或索引读取成绩列。
# 	•	计算平均值。

import pandas as pd

# 读取 Excel 文件
df = pd.read_excel('scores.xlsx')

# 计算平均分
average_score = df['Score'].mean()

# 输出平均分
print(f"Average score: {average_score}")

# 任务：给定一个 Excel 文件 sales_data.xlsx，其中包含每月的销售数据。请在文件中创建一个折线图，显示每个月的销售额。
#
# 思路：
#
# 	•	使用 openpyxl 和 openpyxl.chart 模块。
# 	•	读取销售数据。
# 	•	创建折线图，将数据源设置为销售数据。

import openpyxl
from openpyxl.chart import LineChart, Reference

# 打开 Excel 文件
workbook = openpyxl.load_workbook('sales_data.xlsx')
sheet = workbook.active

# 创建折线图对象
chart = LineChart()

# 设置数据源 (假设销售数据在 B 列，1 行到 12 行)
data = Reference(sheet, min_col=2, min_row=1, max_row=12)

# 添加数据到图表
chart.add_data(data, titles_from_data=True)

# 插入图表到工作表中
sheet.add_chart(chart, 'E5')

# 保存 Excel 文件
workbook.save('sales_data_with_chart.xlsx')

# Excel 文件的筛选和导出
#
# 任务：给定一个 Excel 文件 employees.xlsx，文件中包含员工的姓名、年龄、职位等信息。请筛选出年龄大于 30 岁的员工，并将结果导出到新的 Excel 文件 filtered_employees.xlsx 中。
#
# 思路：
#
# 	•	使用 pandas 读取 Excel 文件。
# 	•	使用筛选条件过滤数据。
# 	•	将过滤后的数据导出到新文件。

import pandas as pd

# 读取 Excel 文件
df = pd.read_excel('employees.xlsx')

# 筛选年龄大于 30 岁的员工
filtered_df = df[df['Age'] > 30]

# 将筛选结果导出到新的 Excel 文件
filtered_df.to_excel('filtered_employees.xlsx', index=False)